"""
Telegram-бот для проверки клиентов по номеру телефона.

Обычный пользователь:
  - присылает номер вида +79261234567
  - бот ищет его в базе и показывает отзыв
  - если номера нет — предлагает добавить (номер + отзыв)

Админ (username из ADMIN_USERNAMES, команды скрыты от остальных):
  /export — выгрузить всю базу в Excel
  /import — загрузить Excel (2 колонки: номер, отзыв), новые номера
            добавляются, уже существующие — не трогаются
  /edit   — изменить отзыв уже существующего клиента
  Каждую ночь в 02:00 бот сам присылает всем админам свежую
  Excel-выгрузку базы в личные сообщения (нужно хотя бы раз
  написать боту, чтобы он узнал chat_id).
"""
import asyncio
import io
import logging
from datetime import datetime, timedelta

import openpyxl
from openpyxl import Workbook

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)

import database as db
from config import ADMIN_USERNAMES, BOT_TOKEN
from phone_utils import normalize_phone

logging.basicConfig(level=logging.INFO)

bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher(storage=MemoryStorage())


class AddClient(StatesGroup):
    waiting_review = State()


class ImportFile(StatesGroup):
    waiting_file = State()


class EditReview(StatesGroup):
    waiting_phone = State()
    waiting_new_review = State()


class DeleteClient(StatesGroup):
    waiting_phone = State()


def _is_admin_username(username: str) -> bool:
    return (username or "").lower() in {a.lower().lstrip("@") for a in ADMIN_USERNAMES}


def is_admin(message: Message) -> bool:
    return _is_admin_username(message.from_user.username)


def is_admin_private(message: Message) -> bool:
    """Админ-команды разрешены ТОЛЬКО в личных сообщениях с ботом —
    никогда в группах, чтобы база не могла случайно утечь в чат
    с посторонними людьми."""
    return message.chat.type == "private" and is_admin(message)


# ---------------------------------------------------------------------------
# Запоминаем chat_id админов при любом их сообщении боту
# (нужно, чтобы бот мог сам написать им в личку для авто-выгрузки)
# ---------------------------------------------------------------------------

@dp.message.middleware()
async def record_admin_chat_id(handler, event: Message, data):
    # Запоминаем chat_id ТОЛЬКО из личных сообщений боту — иначе, если
    # админ написал что-то в группе, где тоже состоит бот, туда же
    # случайно улетит ночная авто-выгрузка базы.
    if event.chat.type == "private" and is_admin(event):
        db.save_admin_chat(event.from_user.username or "", event.chat.id)
    return await handler(event, data)


# ---------------------------------------------------------------------------
# Старт
# ---------------------------------------------------------------------------

@dp.message(CommandStart())
async def cmd_start(message: Message):
    await message.answer(
        "Привет! Пришлите номер телефона в формате +79261234567, "
        "и я проверю, был ли этот клиент у нас в базе."
    )


# ---------------------------------------------------------------------------
# Поиск / добавление клиента
# ---------------------------------------------------------------------------

@dp.message(StateFilter(None), F.text.startswith("+"))
async def handle_phone(message: Message, state: FSMContext):
    phone = normalize_phone(message.text)
    if not phone:
        await message.answer(
            "Не похоже на корректный номер. Пришлите номер в формате +79261234567."
        )
        return

    client = db.find_client(phone)
    if client:
        _, review, _added_by, _added_at, _edited_by, _edited_at = client
        await message.answer(f"📱 Номер: {phone}\n📝 Отзыв: {review}")
    else:
        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="Да, добавить", callback_data=f"add_yes:{phone}"
                    ),
                    InlineKeyboardButton(text="Нет", callback_data="add_no"),
                ]
            ]
        )
        await message.answer(
            "Такого клиента нету, хотите его добавить?", reply_markup=kb
        )


@dp.callback_query(F.data.startswith("add_yes:"))
async def cb_add_yes(callback: CallbackQuery, state: FSMContext):
    phone = callback.data.split(":", 1)[1]
    await state.update_data(phone=phone)
    await state.set_state(AddClient.waiting_review)
    await callback.message.edit_text(f"Введите отзыв о клиенте {phone}:")
    await callback.answer()


@dp.callback_query(F.data == "add_no")
async def cb_add_no(callback: CallbackQuery):
    await callback.message.edit_text("Хорошо, не добавляю.")
    await callback.answer()


@dp.message(AddClient.waiting_review)
async def process_review(message: Message, state: FSMContext):
    data = await state.get_data()
    phone = data.get("phone")
    review = (message.text or "").strip()

    if not review:
        await message.answer("Отзыв не может быть пустым. Введите текст отзыва:")
        return

    if db.phone_exists(phone):
        await message.answer("Этот номер уже кто-то успел добавить, отзыв не изменён.")
    else:
        username = message.from_user.username or f"id{message.from_user.id}"
        db.add_client(phone, review, username)
        await message.answer(f"Готово! Клиент {phone} добавлен в базу.")

    await state.clear()


# ---------------------------------------------------------------------------
# Скрытые админ-команды (не отображаются в меню бота, доступны только
# пользователям с username из ADMIN_USERNAMES; для остальных — тишина)
# ---------------------------------------------------------------------------

def build_export_file() -> BufferedInputFile:
    rows = db.get_all_clients()
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(
        [
            "Номер",
            "Отзыв",
            "Кто добавил",
            "Дата и время добавления",
            "Кто изменил",
            "Дата и время изменения",
        ]
    )
    for phone, review, added_by, added_at, edited_by, edited_at in rows:
        ws.append([phone, review, added_by, added_at, edited_by or "", edited_at or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clients_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return BufferedInputFile(buf.read(), filename=filename)


@dp.message(Command("export"))
async def cmd_export(message: Message):
    if not is_admin_private(message):
        return
    await message.answer_document(build_export_file())


@dp.message(Command("import"))
async def cmd_import(message: Message, state: FSMContext):
    if not is_admin_private(message):
        return

    await state.set_state(ImportFile.waiting_file)
    await message.answer(
        "Пришлите Excel-файл (.xlsx) с двумя колонками: Номер и Отзыв."
    )


@dp.message(ImportFile.waiting_file, F.document)
async def process_import_file(message: Message, state: FSMContext):
    if not is_admin_private(message):
        await state.clear()
        return

    doc = message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.answer("Нужен файл в формате .xlsx. Попробуйте снова: /import")
        await state.clear()
        return

    tg_file = await bot.get_file(doc.file_id)
    file_bytes = await bot.download_file(tg_file.file_path)

    wb = openpyxl.load_workbook(file_bytes)
    ws = wb.active

    added = 0
    skipped = 0
    invalid = 0

    admin_username = message.from_user.username or f"id{message.from_user.id}"

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue

        raw_phone = row[0]
        raw_review = row[1] if len(row) > 1 else None

        phone = normalize_phone(raw_phone)
        review = str(raw_review).strip() if raw_review is not None else ""

        if not phone or not review:
            invalid += 1
            continue

        if db.phone_exists(phone):
            skipped += 1
            continue

        db.add_client(phone, review, f"import:{admin_username}")
        added += 1

    await message.answer(
        "Импорт завершён.\n"
        f"Добавлено новых: {added}\n"
        f"Пропущено (уже были в базе): {skipped}\n"
        f"Некорректных строк (пропущены, например заголовок таблицы): {invalid}"
    )
    await state.clear()


@dp.message(Command("edit"))
async def cmd_edit(message: Message, state: FSMContext):
    if not is_admin_private(message):
        return

    await state.set_state(EditReview.waiting_phone)
    await message.answer("Введите номер телефона клиента, чей отзыв нужно изменить:")


@dp.message(EditReview.waiting_phone)
async def process_edit_phone(message: Message, state: FSMContext):
    if not is_admin_private(message):
        await state.clear()
        return

    phone = normalize_phone(message.text)
    if not phone:
        await message.answer(
            "Не похоже на корректный номер. Пришлите номер в формате +79261234567, "
            "или /start чтобы отменить."
        )
        return

    client = db.find_client(phone)
    if not client:
        await message.answer(
            f"Клиент {phone} не найден в базе. Проверьте номер и попробуйте снова, "
            "или /start чтобы отменить."
        )
        return

    _, current_review, *_ = client
    await state.update_data(phone=phone)
    await state.set_state(EditReview.waiting_new_review)
    await message.answer(
        f"Текущий отзыв о клиенте {phone}:\n«{current_review}»\n\n"
        "Пришлите новый текст отзыва:"
    )


@dp.message(EditReview.waiting_new_review)
async def process_edit_new_review(message: Message, state: FSMContext):
    if not is_admin_private(message):
        await state.clear()
        return

    data = await state.get_data()
    phone = data.get("phone")
    new_review = (message.text or "").strip()

    if not new_review:
        await message.answer("Отзыв не может быть пустым. Введите текст отзыва:")
        return

    editor = message.from_user.username or f"id{message.from_user.id}"
    updated = db.update_review(phone, new_review, editor)

    if updated:
        await message.answer(f"Готово! Отзыв о клиенте {phone} обновлён.")
    else:
        await message.answer(f"Не удалось обновить — клиент {phone} не найден.")

    await state.clear()


@dp.message(Command("delete"))
async def cmd_delete(message: Message, state: FSMContext):
    if not is_admin_private(message):
        return

    await state.set_state(DeleteClient.waiting_phone)
    await message.answer("Введите номер телефона клиента, которого нужно удалить:")


@dp.message(DeleteClient.waiting_phone)
async def process_delete_phone(message: Message, state: FSMContext):
    if not is_admin_private(message):
        await state.clear()
        return

    phone = normalize_phone(message.text)
    if not phone:
        await message.answer(
            "Не похоже на корректный номер. Пришлите номер в формате +79261234567, "
            "или /start чтобы отменить."
        )
        return

    client = db.find_client(phone)
    if not client:
        await message.answer(
            f"Клиент {phone} не найден в базе. Проверьте номер и попробуйте снова, "
            "или /start чтобы отменить."
        )
        return

    _, review, *_ = client
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Да, удалить", callback_data=f"del_yes:{phone}"
                ),
                InlineKeyboardButton(text="Отмена", callback_data="del_no"),
            ]
        ]
    )
    await message.answer(
        f"Удалить клиента {phone}?\nТекущий отзыв: «{review}»",
        reply_markup=kb,
    )
    await state.clear()


@dp.callback_query(F.data.startswith("del_yes:"))
async def cb_delete_yes(callback: CallbackQuery):
    if not _is_admin_username(callback.from_user.username):
        await callback.answer()
        return

    phone = callback.data.split(":", 1)[1]
    deleted = db.delete_client(phone)
    if deleted:
        await callback.message.edit_text(f"Клиент {phone} удалён из базы.")
    else:
        await callback.message.edit_text(f"Клиент {phone} уже отсутствует в базе.")
    await callback.answer()


@dp.callback_query(F.data == "del_no")
async def cb_delete_no(callback: CallbackQuery):
    await callback.message.edit_text("Удаление отменено.")
    await callback.answer()


# ---------------------------------------------------------------------------
# Авто-выгрузка базы всем админам каждую ночь в 02:00
# ---------------------------------------------------------------------------

async def send_nightly_export():
    chat_ids = db.get_admin_chat_ids(list(ADMIN_USERNAMES))
    if not chat_ids:
        logging.warning(
            "Авто-выгрузка: ни один админ ещё ни разу не писал боту, "
            "chat_id неизвестны — отправка пропущена."
        )
        return

    for chat_id in chat_ids:
        try:
            await bot.send_document(
                chat_id,
                build_export_file(),
                caption="Автоматическая ночная выгрузка базы (02:00).",
            )
        except Exception:
            logging.exception("Не удалось отправить авто-выгрузку в chat_id=%s", chat_id)


async def nightly_export_scheduler():
    while True:
        now = datetime.now()
        next_run = now.replace(hour=2, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        await asyncio.sleep((next_run - now).total_seconds())
        await send_nightly_export()


# ---------------------------------------------------------------------------

async def main():
    db.init_db()
    asyncio.create_task(nightly_export_scheduler())
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
